#!/usr/bin/env python3
"""Export and optionally sync the system rules dictionary to Feishu.

The source of truth is config/system_rules.yaml. It uses JSON syntax so the
standard library can parse it without adding a YAML dependency.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import time
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ModuleNotFoundError:  # CSV export still works with the system Python.
    Workbook = None
    Font = PatternFill = Alignment = None

import push_to_feishu as feishu
from feishu_table_registry import table_name


ROOT = Path(__file__).resolve().parents[1]
RULES_PATH = ROOT / "config" / "system_rules.yaml"
OUT = ROOT / "output"
DICTIONARY_TABLE = "99 规则与字典"
CONSOLE_TABLE = "00 主控台"
CONSOLE_CARD_TITLE = "99 规则与字典入口"


def load_rules() -> dict[str, Any]:
    return json.loads(RULES_PATH.read_text(encoding="utf-8"))


def empty_record(fields: list[str], last_updated: str) -> dict[str, str]:
    return {field: "" for field in fields} | {"最后更新时间": last_updated}


def append_record(records: list[dict[str, str]], fields: list[str], last_updated: str, **values: Any) -> None:
    row = empty_record(fields, last_updated)
    for key, value in values.items():
        if key in row:
            row[key] = "" if value is None else str(value)
    records.append(row)


def build_records(rules: dict[str, Any]) -> list[dict[str, str]]:
    fields = rules["dictionary_fields"]
    updated = rules["last_updated"]
    records: list[dict[str, str]] = []

    for table in rules["tables"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="表逻辑",
            所属表=table["name"],
            名称=table["role"],
            说明=f"{table['why']} 定位：{table['role']}",
            输入来源=table["input"],
            输出去向=table["output"],
            生产者=table["producer"],
            消费者=table["consumer"],
            是否用户日常可见=table["daily_visible"],
            是否用户可编辑=table["user_editable"],
            默认视图=table["default_view"],
            AI是否参与="视具体表而定",
            人工判断点=table["human_edit"],
            禁止事项="不要把后台表当每日入口；不要把规则表当业务数据表。",
            备注=f"脚本写入：{table['scripts_write']}",
        )

    for field in rules["fields"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="字段字典",
            所属表=field["table"],
            所属字段=field["field"],
            名称=field["field"],
            说明=field["purpose"],
            生产者=field["producer"],
            消费者=field["consumer"],
            是否用户日常可见="视所在表而定",
            是否用户可编辑=field["user_editable"],
            AI是否参与=field["ai"],
            人工判断点=field["human_judgment"],
            备注=f"必填：{field['required']}",
        )

    for status in rules["status_flows"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="状态字典",
            所属表=status["table"],
            所属字段=status["field"],
            名称=f"{status['flow']} / {status['status']}",
            说明=status["meaning"],
            进入条件=status["entry"],
            退出动作=status["next_action"],
            AI是否参与="可建议，不能替代最终判断",
            人工判断点=f"是否需要人工判断：{status['human_required']}",
            禁止事项="不要把不同状态流混用。",
            **{"状态/取值": status["status"]},
        )

    for scoring in rules["scoring_rules"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="评分规则",
            所属表=table_name("topic_decision"),
            所属字段="总分",
            名称=scoring["dimension"],
            说明=f"高分标准：{scoring['high']} 低分标准：{scoring['low']} 例子：{scoring['example']}",
            评分权重=scoring["weight"],
            AI是否参与="可参与初评",
            人工判断点=f"是否可以人工修正：{scoring['manual_editable']}",
            禁止事项="分数只用于排序，不替代用户主编判断；不要为了追热点牺牲账号定位。",
        )

    for ai_rule in rules["ai_rules"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="AI处理规则",
            名称=ai_rule["name"],
            说明=ai_rule["rule"],
            AI是否参与=f"允许：{ai_rule['allowed']}",
            人工判断点=ai_rule["human_check"],
            禁止事项="不能自动发布、不能伪造数据、不能绕过登录/验证码/反爬、不能生成完整成稿。",
        )

    for console_rule in rules["console_daily_rules"]:
        append_record(
            records,
            fields,
            updated,
            规则类型="主控台/日报规则",
            所属表="00 主控台",
            名称=console_rule["name"],
            说明=console_rule["rule"],
            输出去向=console_rule["output"],
            生产者="规则文件 / 后续脚本",
            消费者="用户本人",
            是否用户日常可见="是",
            是否用户可编辑="是",
            AI是否参与="日报草稿可参与，主控台入口不依赖 AI",
            人工判断点="主控台和日报只提示今天该做什么，最终选择由用户确认。",
            禁止事项="不要把日报写成长报告；不要把主控台扩成新业务流程表。",
        )

    return records


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    if Workbook is None:
        return
    wb = Workbook()
    ws = wb.active
    ws.title = DICTIONARY_TABLE
    headers = list(rows[0].keys())
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in rows:
        ws.append([row.get(header, "") for header in headers])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(42, max(12, max(len(str(c.value or "")) for c in col[:80]) + 2))
    ws.freeze_panes = "A2"
    wb.save(path)


def export_files(rows: list[dict[str, str]]) -> tuple[Path, Path]:
    csv_path = OUT / "system_rules_dictionary.csv"
    xlsx_path = OUT / "system_rules_dictionary.xlsx"
    write_csv(csv_path, rows)
    write_xlsx(xlsx_path, rows)
    return csv_path, xlsx_path


def list_records(token: str, app_token: str, table_id: str) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    page_token = ""
    while True:
        suffix = f"?page_size=500{('&page_token=' + page_token) if page_token else ''}"
        payload = feishu.request_json("GET", f"/bitable/v1/apps/{app_token}/tables/{table_id}/records{suffix}", token=token)
        data = payload.get("data", {})
        items.extend(data.get("items", []))
        if not data.get("has_more"):
            return items
        page_token = data.get("page_token", "")


def create_dictionary_table(token: str, app_token: str, headers: list[str]) -> str:
    payload = feishu.request_json(
        "POST",
        f"/bitable/v1/apps/{app_token}/tables",
        token=token,
        body={
            "table": {
                "name": DICTIONARY_TABLE,
                "default_view_name": "全部",
                "fields": [{"field_name": header, "type": 1} for header in headers],
            }
        },
    )
    return payload.get("data", {}).get("table", {}).get("table_id") or payload.get("data", {}).get("table_id")


def delete_records(token: str, app_token: str, table_id: str, record_ids: list[str]) -> None:
    for start in range(0, len(record_ids), 500):
        chunk = record_ids[start:start + 500]
        feishu.request_json(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_delete",
            token=token,
            body={"records": chunk},
        )
        time.sleep(0.15)


def batch_create(token: str, app_token: str, table_id: str, rows: list[dict[str, str]]) -> None:
    for start in range(0, len(rows), 500):
        chunk = rows[start:start + 500]
        feishu.request_json(
            "POST",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/batch_create",
            token=token,
            body={"records": [{"fields": row} for row in chunk]},
        )
        time.sleep(0.15)


def find_table(token: str, app_token: str, name: str) -> dict[str, Any] | None:
    for table in feishu.list_tables(token, app_token):
        if table.get("name") == name:
            return table
    return None


def ensure_console_card(token: str, app_token: str) -> str:
    table = find_table(token, app_token, CONSOLE_TABLE)
    if not table:
        return "console_missing"
    table_id = table["table_id"]
    records = list_records(token, app_token, table_id)
    fields = {
        "动作": CONSOLE_CARD_TITLE,
        "优先级": "高",
        "工作区": "规则说明",
        "状态": "固定入口",
        "说明": "如果想理解表逻辑、字段含义、状态流转、评分规则、AI边界，请查看 99 规则与字典。它是系统说明书，不是业务数据表。",
        "下一步": "打开 99 规则与字典，按规则类型查看：表逻辑、字段字典、状态字典、评分规则、AI处理规则、主控台/日报规则。",
    }
    existing = next((record for record in records if record.get("fields", {}).get("动作") == CONSOLE_CARD_TITLE), None)
    if existing:
        feishu.request_json(
            "PUT",
            f"/bitable/v1/apps/{app_token}/tables/{table_id}/records/{existing['record_id']}",
            token=token,
            body={"fields": fields},
        )
        return "console_card_updated"
    feishu.request_json(
        "POST",
        f"/bitable/v1/apps/{app_token}/tables/{table_id}/records",
        token=token,
        body={"fields": fields},
    )
    return "console_card_created"


def sync_feishu(rows: list[dict[str, str]]) -> dict[str, Any]:
    token = feishu.tenant_token()
    app_token = feishu.get_or_create_base(token)
    headers = list(rows[0].keys())
    table = find_table(token, app_token, DICTIONARY_TABLE)
    created = False
    if table:
        table_id = table["table_id"]
        existing = list_records(token, app_token, table_id)
        delete_records(token, app_token, table_id, [record["record_id"] for record in existing])
    else:
        table_id = create_dictionary_table(token, app_token, headers)
        created = True
    batch_create(token, app_token, table_id, rows)
    console_status = ensure_console_card(token, app_token)
    return {
        "ok": True,
        "app_token": app_token,
        "table_name": DICTIONARY_TABLE,
        "table_id": table_id,
        "created": created,
        "records": len(rows),
        "console_status": console_status,
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sync-feishu", action="store_true", help="Write 99 规则与字典 and console entry to Feishu.")
    args = parser.parse_args()

    rules = load_rules()
    rows = build_records(rules)
    csv_path, xlsx_path = export_files(rows)
    result: dict[str, Any] = {
        "ok": True,
        "records": len(rows),
        "csv": str(csv_path),
        "xlsx": str(xlsx_path),
    }
    if args.sync_feishu:
        if not (os.getenv("FEISHU_APP_ID") and os.getenv("FEISHU_APP_SECRET") and os.getenv("FEISHU_BASE_APP_TOKEN")):
            raise SystemExit("FEISHU_APP_ID, FEISHU_APP_SECRET and FEISHU_BASE_APP_TOKEN are required for --sync-feishu")
        result["feishu"] = sync_feishu(rows)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
