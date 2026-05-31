#!/usr/bin/env python3
"""
AI账号信息雷达 + 飞书执行台 v0.1

Reads sources.example.yaml, fetches safe public AIHOT data, ingests manual JSONL,
generates explainable analysis CSVs and one XLSX workbook ready for Feishu import.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.error import URLError, HTTPError
from urllib.request import Request, urlopen

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_UA = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 "
    "Safari/537.36 aihot-skill/0.2.0"
)

BUSINESS_KEYWORDS = {
    "content": ["内容", "脚本", "短视频", "公众号", "小红书", "素材", "选题", "写作", "视频", "设计", "营销资产"],
    "marketing": ["营销", "品牌", "增长", "获客", "转化", "投放", "客户", "私域", "销售", "lead", "CRM"],
    "agent": ["Agent", "智能体", "Claude Code", "Codex", "MCP", "workflow", "工作流", "自动化", "子智能体"],
    "video": ["视频", "镜头", "分镜", "导演", "音效", "图像", "AIGC", "成片", "剪辑", "Sora"],
    "process": ["流程", "SOP", "清单", "模板", "表格", "复盘", "交付", "验收", "系统"],
}

PLATFORM_HINTS = {
    "短视频": ["冲突", "对比", "演示", "镜头", "3秒", "误区", "为什么"],
    "小红书": ["清单", "模板", "流程图", "避坑", "收藏", "表格", "工具栈"],
    "公众号": ["框架", "深度", "复盘", "案例", "业务", "方法论", "系统"],
    "X": ["发布", "模型", "开源", "API", "Agent", "论文"],
}


@dataclass
class Source:
    id: str = ""
    source_type: str = ""
    platform: str = ""
    name: str = ""
    homepage_url: str = ""
    fetch_method: str = ""
    feed_url: str = ""
    priority: str = ""
    frequency: str = ""
    focus: str = ""
    notes: str = ""


def parse_sources(path: Path) -> list[Source]:
    """Tiny YAML subset parser for the checked-in sources.example.yaml."""
    sources: list[dict[str, str]] = []
    current: dict[str, str] | None = None
    last_key: str | None = None
    for raw in path.read_text(encoding="utf-8").splitlines():
        line = raw.rstrip()
        stripped = line.strip()
        if not stripped or stripped.startswith("#") or stripped == "sources:":
            continue
        if stripped.startswith("- "):
            if current:
                sources.append(current)
            current = {}
            item = stripped[2:]
            if ":" in item:
                key, value = item.split(":", 1)
                current[key.strip()] = value.strip()
                last_key = key.strip()
            continue
        if current is None:
            continue
        if ":" in stripped:
            key, value = stripped.split(":", 1)
            current[key.strip()] = value.strip()
            last_key = key.strip()
        elif last_key:
            current[last_key] = (current.get(last_key, "") + " " + stripped).strip()
    if current:
        sources.append(current)
    return [Source(**{k: v for k, v in s.items() if k in Source.__annotations__}) for s in sources]


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


def fingerprint(*parts: str) -> str:
    joined = "|".join(p or "" for p in parts)
    return hashlib.sha1(joined.encode("utf-8")).hexdigest()[:16]


def fetch_json(url: str) -> tuple[dict[str, Any] | None, str]:
    req = Request(url, headers={"User-Agent": DEFAULT_UA})
    try:
        with urlopen(req, timeout=20) as resp:
            data = json.loads(resp.read().decode("utf-8"))
            return data, "ok"
    except HTTPError as exc:
        return None, f"http_{exc.code}"
    except (URLError, TimeoutError, json.JSONDecodeError) as exc:
        return None, f"failed:{exc.__class__.__name__}"


def source_config_rows(sources: list[Source]) -> list[dict[str, Any]]:
    return [
        {
            "来源ID": s.id,
            "来源类型": s.source_type,
            "平台": s.platform,
            "来源名称": s.name,
            "主页链接": s.homepage_url,
            "抓取方式": s.fetch_method,
            "RSS/API/网页链接": s.feed_url,
            "是否重点跟踪": "是" if s.priority == "high" else "否",
            "跟踪频率": s.frequency,
            "关注重点": s.focus,
            "备注": s.notes,
        }
        for s in sources
    ]


def load_manual_items(path: Path) -> list[dict[str, Any]]:
    rows = []
    if not path.exists():
        return rows
    for line_no, line in enumerate(path.read_text(encoding="utf-8").splitlines(), start=1):
        if not line.strip():
            continue
        try:
            rows.append(json.loads(line))
        except json.JSONDecodeError as exc:
            print(f"[warn] manual jsonl line {line_no} skipped: {exc}", file=sys.stderr)
    return rows


def build_inbox(sources: list[Source], manual_path: Path, fetch_aihot: bool) -> tuple[list[dict[str, Any]], list[str]]:
    rows: list[dict[str, Any]] = []
    logs: list[str] = []
    seen: set[str] = set()

    if fetch_aihot:
        for source in sources:
            if source.fetch_method not in {"aihot_api", "aihot_daily_api"} or not source.feed_url:
                continue
            data, status = fetch_json(source.feed_url)
            logs.append(f"{source.name}: {status}")
            if not data:
                continue
            if source.fetch_method == "aihot_daily_api":
                for section in data.get("sections", []):
                    for item in section.get("items", []):
                        rows.append(normalize_item(
                            source_type="ai_hotspot",
                            source_name=source.name,
                            platform="AIHOT",
                            title=item.get("title", ""),
                            url=item.get("sourceUrl", ""),
                            published_at=data.get("date", ""),
                            raw_summary=item.get("summary", ""),
                            author=item.get("sourceName", ""),
                            body_snippet=section.get("label", ""),
                        ))
            else:
                for item in data.get("items", []):
                    rows.append(normalize_item(
                        source_type="ai_hotspot",
                        source_name=source.name,
                        platform="AIHOT",
                        title=item.get("title", ""),
                        url=item.get("url", ""),
                        published_at=item.get("publishedAt", ""),
                        raw_summary=item.get("summary", "") or "",
                        author=item.get("source", ""),
                        body_snippet=item.get("category", "") or "",
                    ))

    for item in load_manual_items(manual_path):
        rows.append(normalize_item(**item))

    deduped = []
    for row in rows:
        fp = row["内容指纹"]
        row["是否重复"] = "是" if fp in seen else "否"
        if fp not in seen:
            deduped.append(row)
            seen.add(fp)
    return deduped, logs


def normalize_item(
    source_type: str = "",
    source_name: str = "",
    platform: str = "",
    title: str = "",
    url: str = "",
    published_at: str = "",
    raw_summary: str = "",
    author: str = "",
    body_snippet: str = "",
    **_: Any,
) -> dict[str, Any]:
    fp = fingerprint(url, title, source_name)
    return {
        "来源类型": source_type,
        "来源名称": source_name,
        "平台": platform,
        "标题": title,
        "链接": url,
        "发布时间": published_at,
        "原始摘要": raw_summary,
        "正文片段": body_snippet,
        "作者/账号": author,
        "抓取时间": now_iso(),
        "内容指纹": fp,
        "是否重复": "否",
        "处理状态": "待分析",
    }


def keyword_score(text: str, keywords: list[str], max_score: int = 5) -> int:
    hits = sum(1 for kw in keywords if kw.lower() in text.lower())
    return min(max_score, hits)


def relevance_tags(text: str) -> list[str]:
    tags = []
    for tag, words in BUSINESS_KEYWORDS.items():
        if any(word.lower() in text.lower() for word in words):
            tags.append(tag)
    return tags


def recommend_platform(text: str) -> str:
    scores = {name: keyword_score(text, words, 4) for name, words in PLATFORM_HINTS.items()}
    best = max(scores, key=scores.get)
    if scores[best] == 0:
        return "公众号"
    return best


def level(total: int) -> str:
    if total >= 82:
        return "A-立即转选题"
    if total >= 68:
        return "B-进入候选"
    if total >= 52:
        return "C-观察"
    return "D-跳过"


def analyze_hotspots(inbox: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in inbox:
        if item["来源类型"] != "ai_hotspot":
            continue
        text = f"{item['标题']} {item['原始摘要']} {item['正文片段']}"
        tags = relevance_tags(text)
        source_score = 5 if any(k in item["作者/账号"] for k in ["OpenAI", "Anthropic", "Google", "Mistral", "Claude"]) else 3
        related = min(5, 1 + len(tags))
        asset = 5 if any(k in text for k in ["工作流", "Agent", "工具", "框架", "MCP", "模板", "安全", "流程"]) else 3
        diff = 5 if any(k in text for k in ["Agent", "Claude Code", "Codex", "视频", "营销", "品牌", "内容"]) else 3
        cost_reverse = 4 if item["原始摘要"] else 3
        total = round(
            12 * 4 / 5 + source_score * 14 / 5 + related * 20 / 5 + diff * 16 / 5
            + asset * 16 / 5 + 10 * 4 / 5 + 12 * cost_reverse / 5
        )
        business_scene = choose_scene(text)
        rows.append({
            "关联内容": item["内容指纹"],
            "热点分类": category_cn(item["正文片段"]),
            "内容摘要": item["原始摘要"][:260],
            "核心事件": item["标题"],
            "涉及公司/产品": extract_entities(text),
            "为什么重要": why_important(text, "hotspot"),
            "和我账号的相关性": "、".join(tags) or "需要人工判断",
            "可切入业务场景": business_scene,
            "适合平台": recommend_platform(text),
            "是否适合做观点": yes_no("判断" in text or "发布" in text or "升级" in text or diff >= 4),
            "是否适合做案例": yes_no("案例" in text or "客户" in text or "企业" in text or business_scene != "AI行业短评"),
            "是否适合做清单/SOP": yes_no(asset >= 4),
            "推荐等级": level(total),
            "总分": total,
            "推荐理由": explain_score(total, tags, asset, diff),
            "风险提示": "AIHOT摘要由模型生成，引用前必须打开原文核对；避免把技术更新直接搬成新闻。",
            "是否转入选题": "是" if total >= 68 else "否",
        })
    return rows


def analyze_competitors(inbox: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for item in inbox:
        if item["来源类型"] != "competitor":
            continue
        text = f"{item['标题']} {item['原始摘要']} {item['正文片段']}"
        tags = relevance_tags(text)
        hook = 5 if any(k in text for k in ["别再", "为什么", "差距", "老板", "不是", "先"]) else 3
        structure = 5 if any(k in text for k in ["流程", "三个", "对比", "清单", "拆成", "brief", "分镜"]) else 3
        demand = 5 if any(k in text for k in ["评论", "问", "私信", "领取", "怎么做"]) else 3
        asset = 5 if any(k in text for k in ["模板", "SOP", "清单", "流程", "表格", "brief", "分镜"]) else 3
        diff = 5 if any(k in text for k in ["业务", "导演", "流程", "交付", "老板"]) else 3
        total = round(hook * 14 / 5 + structure * 16 / 5 + demand * 14 / 5 + diff * 18 / 5 + asset * 16 / 5 + 12 * 4 / 5 + 10 * 4 / 5)
        rows.append({
            "关联内容": item["内容指纹"],
            "博主账号": item["作者/账号"] or item["来源名称"],
            "内容摘要": item["原始摘要"],
            "开头钩子": infer_hook(item["标题"]),
            "核心观点": infer_core_view(text),
            "内容结构": infer_structure(text),
            "可学习点": learnable_point(text),
            "不能照搬点": "不能复述对方案例、数据和人设证明；必须换成你的真实业务现场、导演判断和可交付资产。",
            "评论/互动信号": infer_interaction(text),
            "适合我的改写方向": f"从{choose_scene(text)}切入，改成旧流程痛点 -> AI介入点 -> 可展示结果 -> 可领取资产。",
            "推荐等级": level(total),
            "总分": total,
            "推荐理由": explain_score(total, tags, asset, diff),
            "是否转入选题": "是" if total >= 68 else "否",
        })
    return rows


def category_cn(raw: str) -> str:
    return {
        "ai-models": "模型发布/更新",
        "ai-products": "产品发布/更新",
        "industry": "行业动态",
        "paper": "论文研究",
        "tip": "技巧与观点",
    }.get(raw, raw or "未分类")


def yes_no(flag: bool) -> str:
    return "是" if flag else "否"


def choose_scene(text: str) -> str:
    candidates = [
        ("内容团队选题到Brief流程", ["内容", "选题", "脚本", "Brief", "写作"]),
        ("AI导演工作流与视频交付", ["视频", "镜头", "分镜", "导演", "音效", "成片"]),
        ("非技术Agent处理重复业务任务", ["Agent", "智能体", "Claude Code", "Codex", "MCP", "工作流"]),
        ("品牌增长与获客承接", ["品牌", "营销", "获客", "客户", "转化", "增长"]),
        ("项目复盘与能力产品化", ["复盘", "创业", "产品", "服务", "咨询"]),
    ]
    for scene, keys in candidates:
        if any(k.lower() in text.lower() for k in keys):
            return scene
    return "AI行业短评"


def extract_entities(text: str) -> str:
    names = ["OpenAI", "Claude", "Anthropic", "Google", "Gemini", "Mistral", "Perplexity", "Replit", "阿里", "小米", "DeepSeek", "Codex", "Sora"]
    hits = [name for name in names if name.lower() in text.lower()]
    return "、".join(dict.fromkeys(hits)) or "待人工补充"


def why_important(text: str, kind: str) -> str:
    if "Agent" in text or "智能体" in text or "Claude Code" in text or "Codex" in text:
        return "它说明AI正在从聊天工具变成可承接复杂任务的流程代理，适合转译给内容团队和创业项目。"
    if "视频" in text or "音效" in text or "图像" in text:
        return "它关系到AI视频从炫技走向商业内容交付，适合用导演视角讲验收标准。"
    if "营销" in text or "品牌" in text or "获客" in text:
        return "它能连接到品牌增长和线索承接，比单纯工具介绍更接近业务结果。"
    if kind == "hotspot":
        return "它提供了一个可以被二次筛选的AI行业变化，需要落到业务现场才值得做。"
    return "它的结构可学习，但必须转成你的业务现场和资产沉淀。"


def explain_score(total: int, tags: list[str], asset: int, diff: int) -> str:
    reasons = []
    if tags:
        reasons.append(f"命中业务相关标签：{'、'.join(tags)}")
    if asset >= 4:
        reasons.append("可沉淀为清单、SOP、流程图或Brief模板")
    if diff >= 4:
        reasons.append("有差异化解读空间，适合从业务现场而非资讯搬运切入")
    if total >= 82:
        reasons.append("总分高，建议当天进入选题候选")
    elif total >= 68:
        reasons.append("值得进入候选库，等待人工判断表达角度")
    else:
        reasons.append("暂不优先，除非能补充真实案例或用户需求")
    return "；".join(reasons)


def infer_hook(title: str) -> str:
    if "别再" in title:
        return "反常识制止型：先指出常见错误"
    if "为什么" in title:
        return "问题追问型：用原因吸引停留"
    if "差距" in title or "对比" in title:
        return "前后对比型：承诺看到差异"
    return "结果/判断先行型"


def infer_core_view(text: str) -> str:
    if "工具" in text and "流程" in text:
        return "AI价值不在工具数量，而在能否改造具体流程。"
    if "导演" in text or "镜头" in text:
        return "AI视频的关键不是生成，而是导演判断和交付标准。"
    return "把AI变化翻译成用户能执行的动作。"


def infer_structure(text: str) -> str:
    if "对比" in text or "差距" in text:
        return "结果对比 -> 问题背景 -> 拆关键动作 -> 边界/CTA"
    if "三个" in text or "3个" in text:
        return "痛点开场 -> 三个动作 -> 每个动作给业务意义 -> CTA"
    if "分镜" in text or "brief" in text.lower():
        return "Brief -> 分镜/镜头 -> 修改点 -> 成片/验收标准"
    return "冲突句 -> 判断 -> 例子 -> 可执行建议"


def learnable_point(text: str) -> str:
    if "老板" in text:
        return "用老板听得懂的结果语言讲AI，不从工具名开始。"
    if "分镜" in text or "镜头" in text:
        return "把视觉内容拆成可验收步骤，强化AI导演人设。"
    return "学习其钩子和结构，不学习其题材表层。"


def infer_interaction(text: str) -> str:
    if "评论" in text or "问" in text:
        return "已出现评论需求，可作为选题验证信号。"
    return "待人工补充播放、收藏、评论、私信和资料包领取数据。"


def make_topics(hotspots: list[dict[str, Any]], competitors: list[dict[str, Any]], inbox: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_fp = {item["内容指纹"]: item for item in inbox}
    rows = []
    selected = [r for r in hotspots if r["是否转入选题"] == "是"] + [r for r in competitors if r["是否转入选题"] == "是"]
    for row in selected:
        fp = row["关联内容"]
        item = by_fp.get(fp, {})
        text = f"{item.get('标题','')} {item.get('原始摘要','')}"
        scene = row.get("可切入业务场景") or choose_scene(text)
        source_type = item.get("来源类型", "")
        title = topic_title(item.get("标题", ""), scene, source_type)
        positioning = 5 if any(k in scene for k in ["内容", "导演", "Agent", "品牌"]) else 3
        business = 5 if scene != "AI行业短评" else 3
        diff = 5 if source_type in {"ai_hotspot", "competitor"} else 3
        asset = 5 if any(k in text for k in ["流程", "工作流", "模板", "清单", "Agent", "分镜"]) else 4
        timeliness = 5 if source_type == "ai_hotspot" else 3
        conversion = 5 if any(k in scene for k in ["内容团队", "品牌", "Agent", "业务"]) else 3
        cost_reverse = 4
        total = round(positioning * 20 / 5 + business * 18 / 5 + diff * 16 / 5 + asset * 16 / 5 + timeliness * 10 / 5 + conversion * 12 / 5 + cost_reverse * 8 / 5)
        rows.append({
            "选题标题": title,
            "来源内容": fp,
            "来源类型": source_type,
            "目标用户": target_user(scene),
            "业务场景": scene,
            "旧流程痛点": old_pain(scene),
            "AI介入点": ai_entry(scene, text),
            "可展示结果": show_result(scene),
            "可沉淀资产": asset_output(scene),
            "适合平台": row.get("适合平台", recommend_platform(text)),
            "制作成本": "中",
            "转化潜力": "高" if total >= 78 else "中",
            "定位匹配分": positioning,
            "业务现场分": business,
            "差异化分": diff,
            "资产沉淀分": asset,
            "时效性分": timeliness,
            "转化潜力分": conversion,
            "制作成本反向分": cost_reverse,
            "总分": total,
            "状态": "待人工确认",
            "计划发布时间": "",
            "备注": row.get("推荐理由", ""),
        })
    return rows


def topic_title(source_title: str, scene: str, source_type: str) -> str:
    if source_type == "ai_hotspot":
        return f"{source_title}：它对{scene}意味着什么"
    return f"从{source_title}拆一个适合我的{scene}"


def target_user(scene: str) -> str:
    if "内容团队" in scene:
        return "内容负责人、品牌运营、IP操盘手"
    if "导演" in scene:
        return "AIGC视频创作者、品牌内容团队、短视频负责人"
    if "Agent" in scene:
        return "非技术创业者、内容团队负责人、业务流程 owner"
    if "品牌" in scene:
        return "品牌方、增长负责人、创业项目主理人"
    return "关注AI落地的创业者和内容从业者"


def old_pain(scene: str) -> str:
    mapping = {
        "内容团队选题到Brief流程": "每天追热点、临时想选题，Brief缺判断，发布后也难复盘。",
        "AI导演工作流与视频交付": "只会堆prompt和展示结果，缺少镜头、节奏、验收标准。",
        "非技术Agent处理重复业务任务": "把Agent当聊天机器人，任务拆不清、输入输出不稳定。",
        "品牌增长与获客承接": "内容有流量但线索承接弱，资料包、私信和咨询路径断裂。",
        "项目复盘与能力产品化": "只记录情绪和动作，没有沉淀模板、案例和服务入口。",
    }
    return mapping.get(scene, "只知道AI发生了什么，不知道它能改造哪段业务流程。")


def ai_entry(scene: str, text: str) -> str:
    if "Agent" in scene:
        return "把任务拆成输入、处理步骤、验收标准和异常处理，由Agent执行可重复部分。"
    if "导演" in scene:
        return "用AI辅助生成素材，但由人负责Brief、分镜、节奏、修改和验收。"
    if "内容团队" in scene:
        return "用AI完成资料筛选、选题评分、Brief草案和资料包承接设计。"
    if "品牌" in scene:
        return "用AI把用户问题、内容资产和线索动作连接成可复盘流程。"
    return "把热点先翻译成业务问题，再判断是否值得进入选题库。"


def show_result(scene: str) -> str:
    if "导演" in scene:
        return "一页Brief + 3个镜头节点 + 修改前后对比"
    if "Agent" in scene:
        return "任务拆解表 + Agent输入输出示例 + 验收清单"
    return "旧流程/新流程对比表 + 可领取清单"


def asset_output(scene: str) -> str:
    if "导演" in scene:
        return "AI视频Brief与分镜验收清单"
    if "Agent" in scene:
        return "非技术Agent任务拆解模板"
    if "内容团队" in scene:
        return "内容选题评分表和Brief模板"
    return "业务流程改造清单"


def make_briefs(topics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows = []
    for topic in topics:
        rows.append({
            "关联选题": topic["选题标题"],
            "一句话核心判断": f"这条不应被当成AI新闻，而应被翻译成{topic['业务场景']}里的一个可执行流程。",
            "目标用户": topic["目标用户"],
            "开头3秒": f"如果你还只是转发这个热点，你会错过它真正能改造的业务环节。",
            "内容结构": "冲突判断 -> 旧流程痛点 -> AI介入点 -> 可展示结果 -> 边界和CTA",
            "必须讲清的3个点": f"1. {topic['旧流程痛点']} 2. {topic['AI介入点']} 3. 交付物是{topic['可沉淀资产']}而不是完整代写。",
            "可用案例": "用你自己的账号启动、RUNBY/MuseIn、内容团队或汽车品牌传播场景补充。",
            "视觉建议": topic["可展示结果"],
            "CTA": "回复关键词领取清单，或预约AI内容流程诊断。",
            "资料包承接方式": topic["可沉淀资产"],
            "人工补充": "补真实案例、数据、截图、个人判断和不能自动化的边界。",
            "状态": "待人工补充",
        })
    return rows


def review_rows(topics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "关联选题": topic["选题标题"],
            "平台": topic["适合平台"],
            "标题": "",
            "发布时间": "",
            "链接": "",
            "播放/阅读": "",
            "点赞": "",
            "收藏": "",
            "评论": "",
            "转发": "",
            "私信数": "",
            "资料包领取数": "",
            "咨询线索数": "",
            "复盘结论": "",
            "下一步动作": "",
        }
        for topic in topics
    ]


def manual_intake_rows() -> list[dict[str, Any]]:
    return [
        {
            "输入时间": "",
            "输入类型": "对标内容粘贴",
            "来源平台": "抖音/小红书/视频号/公众号/AIHOT/其他",
            "来源名称": "",
            "原始链接": "",
            "原始内容/截图OCR": "把你看到的对标内容、AIHOT条目、截图转文字或链接说明贴这里；后续脚本会解析进内容收件箱。",
            "期望处理": "分析是否值得参考并转选题",
            "处理状态": "待处理",
            "进入内容收件箱": "否",
            "处理备注": "这是最低手工入口：只粘贴原始材料，不需要手动填完整字段。",
        }
    ]


def asset_rows() -> list[dict[str, Any]]:
    return [
        {
            "资产名称": "AI内容业务导演工作流清单",
            "资产类型": "资料包/Lead Magnet",
            "服务的定位假设": "AI业务系统导演型个人IP",
            "服务的选题栏目": "真实工作流改造",
            "目标用户": "内容团队负责人、品牌运营、创业项目主理人",
            "解决的旧流程痛点": "AI工具和热点很多，但团队不知道如何放进选题、脚本、发布和复盘流程。",
            "包含内容": "选题评分表、Brief模板、发布复盘字段、AI介入点清单",
            "关联选题方向": "一个内容团队如何用AI重做选题到Brief流程",
            "承接CTA": "回复关键词领取清单 / 预约AI内容流程诊断",
            "当前状态": "待制作精简版",
            "优先级": "高",
            "备注": "首月最重要资产之一，先做轻量版，不追求完美。",
        },
        {
            "资产名称": "非技术Agent任务拆解模板",
            "资产类型": "模板",
            "服务的定位假设": "把Agent从技术热词翻译成业务任务",
            "服务的选题栏目": "非技术Agent实战",
            "目标用户": "非技术创业者、内容团队负责人、业务流程owner",
            "解决的旧流程痛点": "把Agent当聊天机器人，任务拆不清、输入输出和验收标准不稳定。",
            "包含内容": "任务边界、输入、处理步骤、输出、验收、异常处理",
            "关联选题方向": "竞品监控、周报、素材库、舆情复盘如何Agent化",
            "承接CTA": "回复关键词Agent模板",
            "当前状态": "待制作",
            "优先级": "中高",
            "备注": "适合第3周Agent栏目使用。",
        },
        {
            "资产名称": "AI视频Brief与分镜验收清单",
            "资产类型": "清单/SOP",
            "服务的定位假设": "懂导演的AI内容系统操盘手",
            "服务的选题栏目": "AI导演工作流",
            "目标用户": "AIGC视频创作者、品牌内容团队、短视频负责人",
            "解决的旧流程痛点": "只会生成画面，缺少brief、分镜、节奏、修改和验收标准。",
            "包含内容": "Brief字段、镜头节点、节奏检查、AI味修正、成片验收",
            "关联选题方向": "AI视频不是prompt，而是镜头、节奏和验收标准",
            "承接CTA": "回复关键词AI导演清单",
            "当前状态": "待制作",
            "优先级": "中高",
            "备注": "强化你的导演差异化。",
        },
    ]


def weekly_review_rows() -> list[dict[str, Any]]:
    return [
        {
            "周次": "第1周",
            "时间范围": "",
            "主验证主题": "AI不是工具，是业务流程",
            "重点栏目": "AI业务定调 / 真实工作流改造",
            "本周发布数": "",
            "最高收藏选题": "",
            "最高私信/评论选题": "",
            "资料包需求信号": "",
            "用户真实问题摘录": "",
            "定位假设验证结论": "待复盘",
            "下周加码方向": "",
            "下周砍掉/降频方向": "",
            "需要调整的评分规则": "",
            "备注": "先看收藏、评论质量和私信，不只看播放。",
        },
        {
            "周次": "第2周",
            "时间范围": "",
            "主验证主题": "AI视频不是炫技，是内容交付系统",
            "重点栏目": "AI导演工作流",
            "本周发布数": "",
            "最高收藏选题": "",
            "最高私信/评论选题": "",
            "资料包需求信号": "",
            "用户真实问题摘录": "",
            "定位假设验证结论": "待复盘",
            "下周加码方向": "",
            "下周砍掉/降频方向": "",
            "需要调整的评分规则": "",
            "备注": "验证用户是否需要完整SOP和验收清单。",
        },
        {
            "周次": "第3周",
            "时间范围": "",
            "主验证主题": "非技术人也能理解的业务Agent",
            "重点栏目": "非技术Agent实战",
            "本周发布数": "",
            "最高收藏选题": "",
            "最高私信/评论选题": "",
            "资料包需求信号": "",
            "用户真实问题摘录": "",
            "定位假设验证结论": "待复盘",
            "下周加码方向": "",
            "下周砍掉/降频方向": "",
            "需要调整的评分规则": "",
            "备注": "验证技术内容是否能被业务人听懂。",
        },
        {
            "周次": "第4周",
            "时间范围": "",
            "主验证主题": "账号不是流量工程，是能力产品化",
            "重点栏目": "AI项目复盘 / 商业验证",
            "本周发布数": "",
            "最高收藏选题": "",
            "最高私信/评论选题": "",
            "资料包需求信号": "",
            "用户真实问题摘录": "",
            "定位假设验证结论": "待复盘",
            "下周加码方向": "",
            "下周砍掉/降频方向": "",
            "需要调整的评分规则": "",
            "备注": "验证诊断咨询和工作流共建是否自然出现。",
        },
    ]


def view_navigation_rows() -> list[dict[str, Any]]:
    return [
        {"表名": "内容收件箱", "视图名": "01 今日新内容", "用途": "每天先看今天进入系统的内容", "建议筛选/排序": "抓取时间=今天；是否重复=否；按抓取时间倒序"},
        {"表名": "内容收件箱", "视图名": "02 待分析", "用途": "只看还没处理的内容", "建议筛选/排序": "处理状态=待分析"},
        {"表名": "内容收件箱", "视图名": "03 AIHOT", "用途": "只看AI热点源", "建议筛选/排序": "来源类型=ai_hotspot"},
        {"表名": "内容收件箱", "视图名": "04 对标内容", "用途": "只看对标账号采样内容", "建议筛选/排序": "来源类型=competitor"},
        {"表名": "热点分析表", "视图名": "01 A级热点", "用途": "今天优先转选题的热点", "建议筛选/排序": "推荐等级包含A；按总分降序"},
        {"表名": "热点分析表", "视图名": "02 可做SOP", "用途": "适合沉淀资产的热点", "建议筛选/排序": "是否适合做清单/SOP=是"},
        {"表名": "对标分析表", "视图名": "01 高价值对标", "用途": "优先学习结构的对标内容", "建议筛选/排序": "推荐等级包含A或B；按总分降序"},
        {"表名": "对标分析表", "视图名": "02 钩子库", "用途": "集中看开头钩子", "建议筛选/排序": "按开头钩子分组"},
        {"表名": "选题候选库", "视图名": "01 A级选题", "用途": "优先进入制作的选题", "建议筛选/排序": "总分>=82；按总分降序"},
        {"表名": "选题候选库", "视图名": "02 本周待做", "用途": "排本周内容", "建议筛选/排序": "状态=待人工确认；按总分降序"},
        {"表名": "选题候选库", "视图名": "03 按平台", "用途": "按短视频/小红书/公众号/X安排制作", "建议筛选/排序": "按适合平台分组"},
        {"表名": "内容Brief表", "视图名": "01 待补Brief", "用途": "需要你补真实案例和判断的Brief", "建议筛选/排序": "状态=待人工补充"},
        {"表名": "内容Brief表", "视图名": "02 可制作", "用途": "已补完可开拍/写作的Brief", "建议筛选/排序": "状态=可制作"},
        {"表名": "手动采样入口表", "视图名": "01 待处理", "用途": "你只粘贴原始内容后，我或脚本处理", "建议筛选/排序": "处理状态=待处理"},
        {"表名": "资产与资料包表", "视图名": "01 高优先级资产", "用途": "先做哪些资料包/模板/清单", "建议筛选/排序": "优先级=高或中高；当前状态!=已完成"},
        {"表名": "周复盘与定位校准表", "视图名": "01 本月四周", "用途": "按周验证定位和栏目", "建议筛选/排序": "按周次升序"},
    ]


def logic_rows() -> list[dict[str, Any]]:
    return [
        {
            "表名": "来源配置表",
            "这张表解决什么问题": "把所有信息源显性化：AIHOT、官方博客、工具榜单、技术社区、手动链接、对标博主池。",
            "输入来自哪里": "sources.example.yaml；对标账号来自《AI账号六大内容栏目对标账号研究报告》。",
            "自动处理逻辑": "按来源类型、平台、抓取方式、优先级、频率和关注重点标准化。高风险平台标记为 manual_sample，不做强抓。",
            "输出到哪里": "作为内容收件箱、手动采样和后续采集脚本的来源池。",
            "你要看的判断点": "哪些是重点跟踪，哪些能自动拉取，哪些只能人工采样，关注重点是否还贴合账号定位。",
            "关联表": "内容收件箱",
            "边界/备注": "不保存账号密码、cookie、token；不绕过登录、验证码或反爬。",
        },
        {
            "表名": "内容收件箱",
            "这张表解决什么问题": "把自动抓取和手动粘贴内容先统一收进一个原始池，避免一上来就写稿。",
            "输入来自哪里": "AIHOT API、AIHOT日报、manual_items JSONL、后续RSS/公开网页/手动粘贴。",
            "自动处理逻辑": "统一字段；用标题+链接+来源生成内容指纹；做基础去重；默认状态为待分析。",
            "输出到哪里": "热点分析表、对标分析表。",
            "你要看的判断点": "标题、摘要、来源、是否重复、处理状态；不要在这里直接决定发布。",
            "关联表": "热点分析表、对标分析表",
            "边界/备注": "AIHOT摘要需回原文核对；短视频平台内容只分析你提供的文本/链接/截图转文字。",
        },
        {
            "表名": "热点分析表",
            "这张表解决什么问题": "把AI热点从新闻变成业务现场判断，决定是否值得进入选题。",
            "输入来自哪里": "内容收件箱中来源类型为 ai_hotspot 的内容。",
            "自动处理逻辑": "按内容/营销/Agent/AI视频/业务流程/资产沉淀关键词打标签；判断业务场景、平台、观点/案例/SOP适配；生成推荐等级和理由。",
            "输出到哪里": "选题候选库。",
            "你要看的判断点": "为什么重要、和我账号的相关性、可切入业务场景、推荐理由、风险提示。",
            "关联表": "内容收件箱、选题候选库",
            "边界/备注": "标准不是热点热不热，而是能否转译成内容团队、品牌增长、AI导演、Agent或创业项目流程。",
        },
        {
            "表名": "对标分析表",
            "这张表解决什么问题": "拆对标账号的钩子、结构、信任感和商业入口，但不复制内容。",
            "输入来自哪里": "内容收件箱中来源类型为 competitor 的手动采样内容。",
            "自动处理逻辑": "识别钩子类型、核心观点、内容结构、可学习点、不能照搬点、互动信号和改写方向。",
            "输出到哪里": "选题候选库。",
            "你要看的判断点": "可学习点、不能照搬点、评论/互动信号、适合我的改写方向。",
            "关联表": "来源配置表、内容收件箱、选题候选库",
            "边界/备注": "对标只学结构和验证信号；必须换成你的真实业务现场和资产沉淀。",
        },
        {
            "表名": "选题候选库",
            "这张表解决什么问题": "把高价值热点/对标内容转成你自己的选题候选，并用可解释评分排序。",
            "输入来自哪里": "热点分析表和对标分析表中推荐等级达到 B 以上的内容。",
            "自动处理逻辑": "套用选题公式：真实业务场景 + AI介入点 + 旧流程痛点 + 可展示结果 + 可带走资产；计算定位匹配、业务现场、差异化、资产沉淀、时效性、转化潜力、制作成本反向分。",
            "输出到哪里": "内容Brief表、发布复盘表。",
            "你要看的判断点": "总分只是排序，核心看业务场景、旧流程痛点、AI介入点、可沉淀资产和备注。",
            "关联表": "热点分析表、对标分析表、内容Brief表、发布复盘表",
            "边界/备注": "不追求全平台日更；优先能沉淀模板、SOP、清单、流程图或案例库的选题。",
        },
        {
            "表名": "内容Brief表",
            "这张表解决什么问题": "把选题变成可制作提纲，但保留你的个人判断和成稿表达。",
            "输入来自哪里": "选题候选库。",
            "自动处理逻辑": "生成一句话核心判断、目标用户、开头3秒、内容结构、必须讲清的3点、视觉建议、CTA和资料包承接方式。",
            "输出到哪里": "人工制作、发布复盘表。",
            "你要看的判断点": "人工补充字段：真实案例、业务现场、截图、个人判断、不能自动化的边界。",
            "关联表": "选题候选库、发布复盘表",
            "边界/备注": "只生成Brief，不生成完整小红书文案、公众号文章或短视频成稿。",
        },
        {
            "表名": "发布复盘表",
            "这张表解决什么问题": "把发布后的数据和复盘沉淀回来，决定下周加码什么、砍掉什么。",
            "输入来自哪里": "选题候选库和人工发布数据。",
            "自动处理逻辑": "预建关联选题、平台和数据字段；后续可加收藏率、互动率、线索转化率公式。",
            "输出到哪里": "下一轮选题判断、资料包/服务页优化。",
            "你要看的判断点": "收藏、评论、私信、资料包领取、咨询线索，比播放量更重要。",
            "关联表": "选题候选库、内容Brief表",
            "边界/备注": "不做自动发布；复盘结论必须人工写，因为这关系到账号判断力。",
        },
        {
            "表名": "评分规则总览",
            "这张表解决什么问题": "解释为什么推荐某条内容，而不是只给黑箱分数。",
            "输入来自哪里": "热点/对标分析规则。",
            "自动处理逻辑": "热点评分：时效性12%、来源可信度14%、业务相关度20%、差异化16%、资产沉淀16%、平台适配10%、制作成本反向12%。对标评分：钩子14%、结构16%、评论需求14%、差异化18%、资产沉淀16%、定位匹配12%、成本反向10%。",
            "输出到哪里": "热点分析表、对标分析表、选题候选库。",
            "你要看的判断点": "推荐理由、可学习点、不能照搬点、业务场景切入，要优先于分数。",
            "关联表": "热点分析表、对标分析表、选题候选库",
            "边界/备注": "当前是规则版 v0.1；后续可接飞书AI字段或模型API升级为深度主编判断。",
        },
        {
            "表名": "手动采样入口表",
            "这张表解决什么问题": "把抖音、小红书、视频号、公众号等不能强抓的平台变成最低成本输入入口。",
            "输入来自哪里": "你看到的链接、截图OCR、复制文本、AIHOT页面内容、对标账号内容。",
            "自动处理逻辑": "当前先作为待处理入口；后续脚本会读取待处理记录并标准化写入内容收件箱。",
            "输出到哪里": "内容收件箱、热点分析表、对标分析表。",
            "你要看的判断点": "只需要粘贴原始内容，不需要填完整分析字段。",
            "关联表": "内容收件箱",
            "边界/备注": "不绕过平台限制，不保存登录态。",
        },
        {
            "表名": "资产与资料包表",
            "这张表解决什么问题": "提前规划首月要沉淀的资料包、模板、SOP和清单，避免内容只留下流水账。",
            "输入来自哪里": "定位报告中的首月资产要求、选题候选库中的可沉淀资产。",
            "自动处理逻辑": "预置首月最关键的3个资产方向，后续可由高分选题自动补充。",
            "输出到哪里": "内容Brief表、发布CTA、后续资料包制作。",
            "你要看的判断点": "哪些资产优先做，服务哪个定位假设，承接哪类选题。",
            "关联表": "选题候选库、内容Brief表",
            "边界/备注": "先做精简版，不追求完美。",
        },
        {
            "表名": "周复盘与定位校准表",
            "这张表解决什么问题": "把发布数据回流到定位假设，判断下周加码什么、砍掉什么。",
            "输入来自哪里": "发布复盘表、定位与选题假设、用户评论/私信/收藏信号。",
            "自动处理逻辑": "预置首月4周验证主题；后续可根据发布复盘数据生成周总结。",
            "输出到哪里": "下一周选题规则、栏目比例、资产优先级。",
            "你要看的判断点": "哪个栏目带来真实问题，哪个只带来热闹。",
            "关联表": "发布复盘表、定位与选题假设、选题候选库",
            "边界/备注": "线索表暂不展开，先看前置验证信号。",
        },
        {
            "表名": "视图导航表",
            "这张表解决什么问题": "告诉你每天应该点哪个视图、看什么、按什么筛选，避免表很多但不知道怎么用。",
            "输入来自哪里": "当前执行台结构和使用流程。",
            "自动处理逻辑": "列出关键视图名称、用途和建议筛选/排序。",
            "输出到哪里": "飞书各表视图配置。",
            "你要看的判断点": "先按视图导航使用，不从全量表里翻。",
            "关联表": "全部业务表",
            "边界/备注": "飞书开放API能创建视图名；复杂筛选若API不可写，则先在导航表中显性说明。",
        },
    ]


def positioning_hypothesis_rows() -> list[dict[str, Any]]:
    return [
        {
            "模块": "当前定位假设",
            "假设内容": "AI业务系统导演型个人IP：懂营销、懂内容、懂导演、正在做AI业务系统的人。",
            "依据来自": "《AI账号定位与首月执行精简报告》核心定位句；用户项目背景描述。",
            "为什么这样假设": "中文AI账号里资讯、工具、提示词已经很拥挤；你的差异点在业务现场、内容交付、导演判断、Agent流程和资产沉淀。",
            "需要验证什么": "用户是否把你当成能指导业务动作的人，而不是只看AI资讯的人。",
            "验证信号": "高质量评论、私信提问、资料包领取、咨询线索、收藏率、用户主动描述自己的业务场景。",
            "如果信号不好怎么调整": "减少抽象定位表达，增加真实案例、流程前后对比和具体岗位场景。",
            "状态": "v0.1假设，待验证",
        },
        {
            "模块": "目标用户假设",
            "假设内容": "内容团队负责人、品牌运营/增长负责人、创业项目主理人、非技术但想用AI重做流程的人。",
            "依据来自": "报告中多次出现内容团队、品牌增长、创业项目、非技术Agent、车企/品牌现场。",
            "为什么这样假设": "这些人不一定关心模型细节，但关心效率、交付、内容质量、线索和业务结果。",
            "需要验证什么": "哪一类人最愿意收藏、私信、领取资料包或预约诊断。",
            "验证信号": "私信身份、评论问题、资料包关键词、咨询表单里的行业/岗位。",
            "如果信号不好怎么调整": "把选题按人群拆得更窄，比如只服务内容团队，或只服务创业者，或先从汽车/品牌增长切入。",
            "状态": "v0.1假设，待验证",
        },
        {
            "模块": "不做什么",
            "假设内容": "不做纯AI新闻搬运、工具合集搬运、提示词大全、只炫技的AI视频、泛汽车评测、技术术语堆砌。",
            "依据来自": "定位报告中的“不做什么”和对标报告中的“必须避开什么”。",
            "为什么这样假设": "这些方向容易有流量，但不一定形成你的差异化信任和咨询/工作流共建机会。",
            "需要验证什么": "排除这些内容后，是否仍能稳定获得有效互动和线索。",
            "验证信号": "低播放但高私信/收藏可以保留；高播放但无收藏/私信/线索要谨慎。",
            "如果信号不好怎么调整": "允许少量热点入口内容，但必须加业务解释和资产承接，不回到搬运。",
            "状态": "边界假设，需坚持但可微调比例",
        },
        {
            "模块": "需要的选题形态",
            "假设内容": "一个真实业务场景 + 一个AI介入点 + 一个旧流程痛点 + 一个可展示结果 + 一个可带走资产。",
            "依据来自": "定位报告第四部分“核心选题公式”。",
            "为什么这样假设": "这个公式能把热点和工具从信息转成流程、模板、SOP、清单、案例库或服务入口。",
            "需要验证什么": "这种选题是否比泛工具/泛新闻更容易带来收藏、私信、资料包领取和咨询。",
            "验证信号": "收藏率、完播/阅读完成度、评论中是否出现“怎么做”“能不能给模板”“我们团队也有这个问题”。",
            "如果信号不好怎么调整": "降低概念密度，增加前后对比、截图、表格、案例和更尖锐的开头。",
            "状态": "核心筛选假设",
        },
        {
            "模块": "六大栏目假设",
            "假设内容": "AI业务定调、真实工作流改造、AI导演工作流、非技术Agent实战、AI汽车与品牌增长、AI项目复盘。",
            "依据来自": "两份PDF报告的栏目结构和首月内容比例。",
            "为什么这样假设": "六个栏目分别承担判断力、主菜流程、视觉护城河、技术转译、行业差异化和活人复盘。",
            "需要验证什么": "哪两个栏目最能带来高质量关注和线索，哪些只是热闹。",
            "验证信号": "按栏目统计收藏、评论、私信、资料包领取、咨询线索。",
            "如果信号不好怎么调整": "首月结束后保留2-3个强栏目，弱栏目降频或并入强栏目。",
            "状态": "栏目组合假设，首月验证",
        },
        {
            "模块": "评分规则来源",
            "假设内容": "热点评分看时效、来源可信、业务相关、差异化、资产沉淀、平台适配、成本；对标评分看钩子、结构、评论需求、差异化、资产沉淀、定位匹配、成本。",
            "依据来自": "用户需求中明确提出的评分维度 + 两份报告中的选题过滤五问和对标拆解方法。",
            "为什么这样假设": "评分不是为了替你判断，而是帮你从大量信息中先排队，降低启动期琐碎筛选成本。",
            "需要验证什么": "高分选题是否真的在发布后带来更好收藏、私信、资料包领取和咨询。",
            "验证信号": "选题总分与发布复盘数据的相关性；高分低反馈要回看规则。",
            "如果信号不好怎么调整": "下调时效/平台适配权重，上调业务现场、资产沉淀、转化潜力或真实案例权重。",
            "状态": "v0.1规则假设",
        },
        {
            "模块": "内容生成边界",
            "假设内容": "系统只生成Brief，不生成完整小红书文案、公众号文章或短视频成稿。",
            "依据来自": "用户明确要求保留个人判断、表达、人设和业务现场感。",
            "为什么这样假设": "账号信任来自你的判断和真实现场，不应让AI替你完成最终表达。",
            "需要验证什么": "Brief是否足够降低启动期琐事，同时不稀释个人表达。",
            "验证信号": "你从Brief到发布的耗时、发布质量、是否仍有个人判断和真实案例。",
            "如果信号不好怎么调整": "Brief增加案例槽位、素材清单和镜头建议，但仍不生成完整成稿。",
            "状态": "明确边界",
        },
        {
            "模块": "商业验证假设",
            "假设内容": "首月先验证咨询和工作流诊断，不急着卖课、卖泛模板或接泛广告。",
            "依据来自": "定位报告商业路径和首月KPI。",
            "为什么这样假设": "你的优势更接近服务和结果交付，先卖诊断/共建能获得真实需求和案例。",
            "需要验证什么": "内容是否能自然引出诊断名额、工作流共建、资料包领取和有效私信。",
            "验证信号": "咨询线索数、资料包领取数、用户问题质量、是否出现付费测试机会。",
            "如果信号不好怎么调整": "优化CTA和资料包，增加服务页式内容，减少泛观点内容。",
            "状态": "首月商业假设",
        },
    ]


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for name, rows in sheets.items():
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws.append(["暂无数据"])
            continue
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for col_idx, header in enumerate(headers, start=1):
            values = [str(r.get(header, "")) for r in rows[:50]]
            width = min(42, max(10, len(header) + 2, *(min(40, len(v)) for v in values)))
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--sources", default=str(ROOT / "sources.example.yaml"))
    parser.add_argument("--manual", default=str(ROOT / "data/manual/manual_items.example.jsonl"))
    parser.add_argument("--out", default=str(ROOT / "output"))
    parser.add_argument("--no-fetch", action="store_true", help="skip live AIHOT fetch and only use manual JSONL")
    args = parser.parse_args()

    out = Path(args.out)
    sources = parse_sources(Path(args.sources))
    inbox, logs = build_inbox(sources, Path(args.manual), fetch_aihot=not args.no_fetch)
    hotspot_rows = analyze_hotspots(inbox)
    competitor_rows = analyze_competitors(inbox)
    topics = make_topics(hotspot_rows, competitor_rows, inbox)
    briefs = make_briefs(topics)
    reviews = review_rows(topics)
    logic = logic_rows()
    positioning = positioning_hypothesis_rows()
    manual_intake = manual_intake_rows()
    assets = asset_rows()
    weekly = weekly_review_rows()
    view_nav = view_navigation_rows()

    sheets = {
        "定位与选题假设": positioning,
        "执行台逻辑说明": logic,
        "视图导航表": view_nav,
        "来源配置表": source_config_rows(sources),
        "手动采样入口表": manual_intake,
        "内容收件箱": inbox,
        "热点分析表": hotspot_rows,
        "对标分析表": competitor_rows,
        "选题候选库": topics,
        "内容Brief表": briefs,
        "资产与资料包表": assets,
        "发布复盘表": reviews,
        "周复盘与定位校准表": weekly,
    }
    names = {
        "定位与选题假设": "positioning_hypothesis.csv",
        "执行台逻辑说明": "execution_logic.csv",
        "视图导航表": "view_navigation.csv",
        "来源配置表": "sources_config.csv",
        "手动采样入口表": "manual_intake.csv",
        "内容收件箱": "content_inbox.csv",
        "热点分析表": "hotspot_analysis.csv",
        "对标分析表": "competitor_analysis.csv",
        "选题候选库": "topic_candidates.csv",
        "内容Brief表": "content_briefs.csv",
        "资产与资料包表": "assets.csv",
        "发布复盘表": "publishing_review.csv",
        "周复盘与定位校准表": "weekly_review.csv",
    }
    for sheet_name, rows in sheets.items():
        write_csv(out / names[sheet_name], rows)
    write_workbook(out / "feishu_import_workbook.xlsx", sheets)
    (out / "run_log.json").write_text(json.dumps({
        "generated_at": now_iso(),
        "aihot_fetch": logs,
        "counts": {k: len(v) for k, v in sheets.items()},
        "note": "AIHOT失败不会中断；可用 --no-fetch 仅跑手动导入。",
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"output": str(out), "counts": {k: len(v) for k, v in sheets.items()}, "aihot_fetch": logs}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
